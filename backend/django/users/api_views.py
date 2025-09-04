from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth.models import User
from django.conf import settings
from .serializers import UserSerializer, UserCreateSerializer
from pathlib import Path
import os
import re
import unicodedata
from datetime import datetime
from django.db import connections

# ====== Helpers de detección de vistas por esquema ======
def _view_exists(schema: str, view: str) -> bool:
    with connections['default'].cursor() as c:
        c.execute(
            """
            SELECT 1 FROM information_schema.views
            WHERE table_schema=%s AND table_name=%s
            """,
            [schema, view],
        )
        return c.fetchone() is not None

def _choose_view(candidates):
    """Recibe lista de (schema, view). Retorna 'schema.view' de la primera que exista o None."""
    for schema, view in candidates:
        try:
            if _view_exists(schema, view):
                return f"{schema}.{view}"
        except Exception:
            continue
    return None
class UserListView(generics.ListCreateAPIView):
    queryset = User.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserCreateSerializer
        return UserSerializer

class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_profile(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([AllowAny])
def api_status(request):
    return Response({
        'status': 'API funcionando correctamente',
        'version': '1.0',
        'mensaje': 'API lista para desarrollo frontend',
        'endpoints': {
            'publicos': [
                '/api/v1/status/',
                '/api/v1/users/ (GET, POST)',
            ],
            'autenticados': [
                '/api/v1/profile/',
                '/api/v1/users/{id}/ (GET, PUT, DELETE)'
            ]
        },
        'autenticacion': {
            'tipo': 'Token',
            'header': 'Authorization: Token tu_token_aqui',
            'ejemplo_token': 'a4c54ef00b2fc005a533c137d153949871f12f22'
        }
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def api_test_data(request):
    """Endpoint de prueba con datos mock para el frontend"""
    return Response({
        'usuarios_ejemplo': [
            {'id': 1, 'nombre': 'Juan Pérez', 'email': 'juan@coofisam.com'},
            {'id': 2, 'nombre': 'María González', 'email': 'maria@coofisam.com'}
        ],
        'configuracion': {
            'cors_habilitado': True,
            'base_url': 'http://IP_SERVIDOR:8000/api/v1/'
        }
    })

# ====== Utilidades módulo financiero ======

def _norm(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c)).lower()

MESES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'setiembre': 9,
    'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}

FNAME_RE = re.compile(
    r'listado[\s_-]*balances[\s_-]*consolidado[\s_-]*([A-Za-zÁÉÍÓÚáéíóúñÑ]+)[\s_-]*(\d{4})\.(xlsx|xls)$',
    re.IGNORECASE
)

def _get_balance_root() -> Path:
    # Preferir carpeta de subidos definida en settings
    root = getattr(settings, 'LIBRO_BALANCE_ROOT', None)
    if root:
        return Path(root)
    # Fallback relativo al proyecto
    return Path(settings.BASE_DIR) / 'Coofisam' / 'data' / 'Libro_de_Balance_subidos'

def _build_tree(path: Path, max_depth: int = 3, include_files: bool = True, _depth: int = 0):
    node = {"name": path.name or str(path), "type": "dir", "children": []}
    if _depth >= max_depth:
        return node
    try:
        with os.scandir(path) as it:
            entries = sorted(it, key=lambda e: (not e.is_dir(follow_symlinks=False), e.name.lower()))
            for entry in entries:
                if entry.is_dir(follow_symlinks=False):
                    node["children"].append(_build_tree(Path(entry.path), max_depth, include_files, _depth + 1))
                elif include_files:
                    node["children"].append({"name": entry.name, "type": "file"})
    except FileNotFoundError:
        node["error"] = f"Ruta no existe: {path}"
    except PermissionError:
        node["children"].append({"name": "[permiso denegado]", "type": "file"})
    return node


# ====== API módulo financiero ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_tree(request):
    """Arbol de carpetas de LIBRO_BALANCE_ROOT (subidos). Params: depth, includeFiles."""
    depth = int(request.query_params.get('depth', 3))
    include_files = request.query_params.get('includeFiles', 'true').lower() != 'false'
    root = _get_balance_root()
    data = _build_tree(root, max_depth=max(1, min(depth, 8)), include_files=include_files)
    return Response({
        'root': root.name,
        'absolute_root': str(root),
        'tree': data.get('children', []),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finanzas_upload(request):
    """Sube archivo Excel: Listado_balances_Consolidado_<Mes>_<Año>.xlsx al directorio Aanoo_<AÑO>."""
    f = request.FILES.get('file')
    if not f:
        return Response({"success": False, "error": "Archivo no encontrado (campo 'file')."}, status=400)

    m = FNAME_RE.search(f.name)
    if not m:
        return Response({
            "success": False,
            "error": "Nombre inválido. Usa: Listado_balances_Consolidado_<Mes>_<Año>.xlsx"
        }, status=400)

    mes_txt, anio_txt = m.group(1), m.group(2)
    if _norm(mes_txt) not in MESES:
        return Response({"success": False, "error": f"Mes no reconocido: {mes_txt}"}, status=400)

    anio = int(anio_txt)
    root = _get_balance_root()
    target_dir = root / f"Aanoo_{anio}"
    target_dir.mkdir(parents=True, exist_ok=True)
    dest_path = target_dir / Path(f.name).name
    if dest_path.exists():
        return Response({"success": False, "error": f"Ya existe un archivo con ese nombre en {target_dir}"}, status=400)

    with open(dest_path, "wb+") as dst:
        for chunk in f.chunks():
            dst.write(chunk)

    return Response({
        "success": True,
        "saved_name": dest_path.name,
        "anio_dir": f"Aanoo_{anio}",
        "dest_path": str(dest_path),
    }, status=201)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_files(request):
    """Lista los archivos subidos, agrupados por año (Aanoo_YYYY)."""
    root = _get_balance_root()
    result = []
    if not root.exists():
        return Response({"files": result})
    for year_dir in sorted(root.glob('Aanoo_*')):
        if year_dir.is_dir():
            year_item = {
                'year': year_dir.name.replace('Aanoo_', ''),
                'path': str(year_dir),
                'files': [p.name for p in sorted(year_dir.glob('*.xls*'))]
            }
            result.append(year_item)
    return Response({"files": result})


@api_view(['GET'])
@permission_classes([AllowAny])
def finanzas_sample(request):
    """Datos mock para desarrollar UI financiera en React."""
    return Response({
        'menus': [
            {'key': 'carga_balance', 'label': 'Carga de Libro de Balance'},
            {'key': 'indicadores', 'label': 'Indicadores'},
            {'key': 'cupos', 'label': 'Cupos'},
        ],
        'upload': {
            'pattern': 'Listado_balances_Consolidado_<Mes>_<Año>.xlsx',
            'months_supported': list(MESES.keys()),
            'destination': 'Aanoo_<AÑO> dentro de LIBRO_BALANCE_ROOT',
        },
        'endpoints': {
            'tree': '/api/v1/finanzas/tree/',
            'upload': '/api/v1/finanzas/upload/',
            'files': '/api/v1/finanzas/files/',
            'indicadores_spec': '/api/v1/finanzas/indicadores/spec/',
            'indicadores_list': '/api/v1/finanzas/indicadores/',
            'indicadores_series': '/api/v1/finanzas/indicadores/series/',
            'cupos_spec': '/api/v1/finanzas/cupos/spec/',
            'cupos_list': '/api/v1/finanzas/cupos/',
            'cupos_credito_spec': '/api/v1/finanzas/cupos-credito/spec/',
            'cupos_credito_list': '/api/v1/finanzas/cupos-credito/',
            'presupuesto_spec': '/api/v1/finanzas/presupuesto/spec/',
            'presupuesto_list': '/api/v1/finanzas/presupuesto/',
            'indicadores_consolidados': '/api/v1/finanzas/indicadores/consolidados/',
        }
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_indicadores_spec(request):
    spec = {
        'filters': [
            {'name': 'year', 'label': 'Año', 'type': 'integer', 'min': 2018, 'max': 2035, 'default': datetime.now().year},
            {'name': 'month', 'label': 'Mes', 'type': 'integer', 'min': 1, 'max': 12},
            {'name': 'agency', 'label': 'Agencia', 'type': 'select', 'options': [
                {'value': 'central', 'label': 'Agencia Central'},
                {'value': 'norte', 'label': 'Agencia Norte'},
                {'value': 'sur', 'label': 'Agencia Sur'},
            ]},
        ],
        'kpis': [
            {
                'id': 'cartera_total', 'name': 'Cartera Total', 'unit': 'COP', 'decimals': 0,
                'direction': 'higher_is_better', 'formula': 'SUM(saldo_cartera)', 'source': 'libro_balance'
            },
            {
                'id': 'mora', 'name': 'Mora (%)', 'unit': '%', 'decimals': 2,
                'direction': 'lower_is_better', 'formula': 'cartera_en_mora / cartera_total * 100', 'source': 'libro_balance'
            },
            {
                'id': 'liquidez', 'name': 'Índice de Liquidez', 'unit': 'ratio', 'decimals': 2,
                'direction': 'higher_is_better', 'formula': 'activos_corrientes / pasivos_corrientes', 'source': 'libro_balance'
            },
        ],
        'table': {
            'columns': [
                {'key': 'period', 'label': 'Periodo', 'type': 'string'},
                {'key': 'value', 'label': 'Valor', 'type': 'number'}
            ]
        }
    }
    return Response(spec)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_indicadores_list(request):
    year = int(request.query_params.get('year', datetime.now().year))
    month = int(request.query_params.get('month', datetime.now().month))
    agency = request.query_params.get('agency', 'central')
    data = [
        {'kpi': 'cartera_total', 'label': 'Cartera Total', 'unit': 'COP', 'value': 12500000000, 'year': year, 'month': month, 'agency': agency},
        {'kpi': 'mora', 'label': 'Mora (%)', 'unit': '%', 'value': 4.25, 'year': year, 'month': month, 'agency': agency},
        {'kpi': 'liquidez', 'label': 'Índice de Liquidez', 'unit': 'ratio', 'value': 1.23, 'year': year, 'month': month, 'agency': agency},
    ]
    return Response({'items': data, 'filters': {'year': year, 'month': month, 'agency': agency}})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_indicadores_series(request):
    kpi = request.query_params.get('kpi', 'cartera_total')
    year = int(request.query_params.get('year', datetime.now().year))
    agency = request.query_params.get('agency', 'central')
    base = {
        'cartera_total': 10000000000,
        'mora': 5.0,
        'liquidez': 1.1,
    }.get(kpi, 0)
    series = []
    for m in range(1, 13):
        if kpi == 'cartera_total':
            val = base * (1 + (m - 6) * 0.01)
        elif kpi == 'mora':
            val = max(2.5, base + (6 - m) * 0.15)
        elif kpi == 'liquidez':
            val = base + (m % 3) * 0.05
        else:
            val = base
        series.append({'month': m, 'value': round(val, 2) if isinstance(val, float) else int(val)})
    return Response({'kpi': kpi, 'year': year, 'agency': agency, 'series': series})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_cupos_spec(request):
    spec = {
        'form': [
            {'name': 'fecha', 'label': 'Fecha', 'type': 'date', 'required': True},
            {'name': 'agencia', 'label': 'Agencia', 'type': 'select', 'required': True, 'options': [
                {'value': 'central', 'label': 'Agencia Central'},
                {'value': 'norte', 'label': 'Agencia Norte'},
                {'value': 'sur', 'label': 'Agencia Sur'},
            ]},
            {'name': 'producto', 'label': 'Producto', 'type': 'select', 'required': True, 'options': [
                {'value': 'consumo', 'label': 'Consumo'},
                {'value': 'microcredito', 'label': 'Microcrédito'},
                {'value': 'comercial', 'label': 'Comercial'},
            ]},
            {'name': 'cupo_asignado', 'label': 'Cupo Asignado', 'type': 'number', 'min': 0, 'required': True},
            {'name': 'cupo_utilizado', 'label': 'Cupo Utilizado', 'type': 'number', 'min': 0, 'required': True},
            {'name': 'observaciones', 'label': 'Observaciones', 'type': 'text', 'maxLength': 500},
        ],
        'constraints': [
            {'rule': 'cupo_utilizado <= cupo_asignado', 'message': 'El cupo utilizado no puede superar el asignado'},
            {'rule': 'cupo_asignado >= 0 and cupo_utilizado >= 0', 'message': 'Valores no pueden ser negativos'},
        ],
        'table': {
            'columns': [
                {'key': 'fecha', 'label': 'Fecha'},
                {'key': 'agencia', 'label': 'Agencia'},
                {'key': 'producto', 'label': 'Producto'},
                {'key': 'cupo_asignado', 'label': 'Cupo Asignado', 'type': 'number'},
                {'key': 'cupo_utilizado', 'label': 'Utilizado', 'type': 'number'},
                {'key': 'cupo_disponible', 'label': 'Disponible', 'type': 'number'},
                {'key': 'observaciones', 'label': 'Observaciones'},
            ]
        }
    }
    return Response(spec)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def finanzas_cupos(request):
    if request.method == 'GET':
        items = [
            {
                'fecha': '2025-08-01', 'agencia': 'central', 'producto': 'consumo',
                'cupo_asignado': 1200000000, 'cupo_utilizado': 450000000,
                'cupo_disponible': 750000000, 'observaciones': 'Campaña mitad de año'
            },
            {
                'fecha': '2025-08-01', 'agencia': 'norte', 'producto': 'microcredito',
                'cupo_asignado': 600000000, 'cupo_utilizado': 420000000,
                'cupo_disponible': 180000000, 'observaciones': ''
            },
        ]
        return Response({'items': items})

    payload = request.data.copy()
    try:
        asignado = float(payload.get('cupo_asignado', 0))
        utilizado = float(payload.get('cupo_utilizado', 0))
    except ValueError:
        return Response({'error': 'Valores numéricos inválidos'}, status=400)

    if asignado < 0 or utilizado < 0:
        return Response({'error': 'Valores no pueden ser negativos'}, status=400)
    if utilizado > asignado:
        return Response({'error': 'El cupo utilizado no puede superar el asignado'}, status=400)

    disponible = asignado - utilizado
    created = {
        'fecha': payload.get('fecha') or datetime.now().strftime('%Y-%m-%d'),
        'agencia': payload.get('agencia', 'central'),
        'producto': payload.get('producto', 'consumo'),
        'cupo_asignado': asignado,
        'cupo_utilizado': utilizado,
        'cupo_disponible': disponible,
        'observaciones': payload.get('observaciones', ''),
        'created_at': datetime.now().isoformat(timespec='seconds')
    }
    return Response(created, status=201)


# ====== Cupos de Crédito ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_cupos_credito_spec(request):
    """Esquema de la Tabla 2: Cupos de Crédito."""
    spec = {
        'columns': [
            {'key': 'fecha_renovado', 'label': 'Fecha Renovado', 'type': 'string'},
            {'key': 'cuenta', 'label': 'Cuenta (Id_cuenta)', 'type': 'string'},
            {'key': 'entidad_financiera', 'label': 'Entidad Financiera', 'type': 'string'},
            {'key': 'cupo_asignado', 'label': 'Cupo Asignado', 'type': 'number'},
            {'key': 'cupo_ejecutado', 'label': 'Cupo Ejecutado', 'type': 'number'},
            {'key': 'disponible', 'label': 'Disponible', 'type': 'number'},
            {'key': 'garantia', 'label': 'Garantía', 'type': 'string'},
            {'key': 'porcentaje_utilizacion', 'label': '% Utilización', 'type': 'number'},
            {'key': 'plazo', 'label': 'Plazo', 'type': 'string'},
        ]
    }
    return Response(spec)


@api_view(['GET'])
@permission_classes([IsAuthenticated])

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def finanzas_cupos_credito_list(request):
    """Cupos de Crédito desde tablas del esquema finanzas (tabla: finanzas.cupos_bancarios).

    GET filtros opcionales: ?year=YYYY&month=M&entidad=...&limit=100
    POST: inserta/actualiza registro. Si se envía id, hace UPDATE; si no, INSERT.
    """
    if request.method == 'POST':
        data = request.data
        rec_id = data.get('id')
        entidad = (data.get('entidad_financiera') or data.get('entidad') or '').strip()
        cuenta = (data.get('cuenta') or data.get('cuentas_balance') or '').strip()
        fecha = (data.get('fecha_renovado') or '').strip()
        def to_float(x):
            try:
                return None if x in (None, '', '-') else float(x)
            except Exception:
                return None
        cupo_asignado = to_float(data.get('cupo_asignado'))
        cupo_ejecutado = to_float(data.get('cupo_ejecutado'))
        garantia = (data.get('garantia') or '').strip() or None
        plazo_meses = data.get('plazo_meses') or data.get('plazo')
        try:
            plazo_meses = int(str(plazo_meses).split()[0]) if plazo_meses not in (None, '') else None
        except Exception:
            plazo_meses = None
        tasa_pct = to_float(data.get('tasa') or data.get('tasa_pct'))

        if not (entidad and fecha and cupo_asignado is not None):
            return Response({'error': 'entidad_financiera, fecha_renovado y cupo_asignado son obligatorios'}, status=400)

        disponible = None
        utilizacion_pct = None
        if cupo_asignado is not None and cupo_ejecutado is not None:
            disponible = cupo_asignado - cupo_ejecutado
            try:
                utilizacion_pct = (cupo_ejecutado / cupo_asignado) * 100 if cupo_asignado else 0
            except Exception:
                utilizacion_pct = None

        from django.db import connections
        with connections['default'].cursor() as c:
            if rec_id:
                c.execute(
                    """
                    UPDATE finanzas.cupos_bancarios
                    SET entidad_financiera=%s, cuentas_balance=%s, fecha_renovado=%s,
                        cupo_asignado=%s, cupo_ejecutado=%s, disponible=COALESCE(%s, cupo_asignado - COALESCE(cupo_ejecutado,0)),
                        garantia=%s, utilizacion_pct=%s, plazo_meses=%s, tasa_pct=%s
                    WHERE id=%s
                    """,
                    [entidad, cuenta or None, fecha, cupo_asignado, cupo_ejecutado, disponible, garantia, utilizacion_pct, plazo_meses, tasa_pct, rec_id]
                )
                return Response({'success': True, 'updated_id': rec_id})
            else:
                c.execute(
                    """
                    INSERT INTO finanzas.cupos_bancarios
                        (entidad_financiera, cuentas_balance, fecha_renovado, cupo_asignado,
                         cupo_ejecutado, disponible, garantia, utilizacion_pct, plazo_meses, tasa_pct, created_at)
                    VALUES (%s,%s,%s,%s,%s,COALESCE(%s, %s - COALESCE(%s,0)),%s,%s,%s,%s, NOW())
                    RETURNING id
                    """,
                    [entidad, cuenta or None, fecha, cupo_asignado, cupo_ejecutado, disponible, cupo_asignado, cupo_ejecutado, garantia, utilizacion_pct, plazo_meses, tasa_pct]
                )
                new_id = c.fetchone()[0]
                return Response({'success': True, 'id': new_id}, status=201)

    # GET
    year = request.query_params.get('year')
    month = request.query_params.get('month')
    entidad = request.query_params.get('entidad')
    limit = int(request.query_params.get('limit', '200'))
    where = []
    params = {}
    if year:
        where.append('EXTRACT(year FROM fecha_renovado) = %(year)s::int')
        params['year'] = year
    if month:
        where.append('EXTRACT(month FROM fecha_renovado) = %(month)s::int')
        params['month'] = month
    if entidad:
        where.append('entidad_financiera ILIKE %(entidad)s')
        params['entidad'] = f'%{entidad}%'
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    sql = f"""
        SELECT 
            fecha_renovado::text AS fecha_renovado,
            cuentas_balance::text AS cuenta,
            entidad_financiera::text AS entidad_financiera,
            cupo_asignado::numeric AS cupo_asignado,
            cupo_ejecutado::numeric AS cupo_ejecutado,
            disponible::numeric AS disponible,
            garantia::text AS garantia,
            utilizacion_pct::numeric AS porcentaje_utilizacion,
            COALESCE(plazo_meses::text || ' meses', NULL) AS plazo,
            tasa_pct::numeric AS tasa
        FROM finanzas.cupos_bancarios
        {where_sql}
        ORDER BY fecha_renovado DESC NULLS LAST, entidad_financiera, cuentas_balance
        LIMIT {limit}
    """
    from django.db import connections
    try:
        with connections['default'].cursor() as c:
            c.execute(sql, params)
            cols = [col[0] for col in c.description]
            items = [dict(zip(cols, row)) for row in c.fetchall()]
        return Response({'source': 'finanzas.cupos_bancarios', 'count': len(items), 'items': items, 'filters': {'year': year, 'month': month, 'entidad': entidad}})
    except Exception as e:
        return Response({'error': str(e), 'source': 'finanzas.cupos_bancarios'}, status=400)


class CuposCreditoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        entidad = request.query_params.get('entidad')
        rec_id = request.query_params.get('id')
        limit = int(request.query_params.get('limit', '200'))
        where = []
        params = {}
        if year:
            where.append('EXTRACT(year FROM fecha_renovado) = %(year)s::int')
            params['year'] = year
        if month:
            where.append('EXTRACT(month FROM fecha_renovado) = %(month)s::int')
            params['month'] = month
        if entidad:
            where.append('entidad_financiera ILIKE %(entidad)s')
            params['entidad'] = f'%{entidad}%'
        if rec_id:
            where.append('id = %(id)s::bigint')
            params['id'] = rec_id
        where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

        sql = f"""
            SELECT 
                id::bigint AS id,
                fecha_renovado::text AS fecha_renovado,
                cuentas_balance::text AS cuenta,
                entidad_financiera::text AS entidad_financiera,
                cupo_asignado::numeric AS cupo_asignado,
                cupo_ejecutado::numeric AS cupo_ejecutado,
                disponible::numeric AS disponible,
                garantia::text AS garantia,
                utilizacion_pct::numeric AS porcentaje_utilizacion,
                COALESCE(plazo_meses::text || ' meses', NULL) AS plazo,
                tasa_pct::numeric AS tasa
            FROM finanzas.cupos_bancarios
            {where_sql}
            ORDER BY fecha_renovado DESC NULLS LAST, entidad_financiera, cuentas_balance
            LIMIT {limit}
        """
        try:
            with connections['default'].cursor() as c:
                c.execute(sql, params)
                cols = [col[0] for col in c.description]
                items = [dict(zip(cols, row)) for row in c.fetchall()]
            return Response({'source': 'finanzas.cupos_bancarios', 'count': len(items), 'items': items, 'filters': {'year': year, 'month': month, 'entidad': entidad}})
        except Exception as e:
            return Response({'error': str(e), 'source': 'finanzas.cupos_bancarios'}, status=400)

    def post(self, request):
        data = request.data
        rec_id = data.get('id')
        entidad = (data.get('entidad_financiera') or data.get('entidad') or '').strip()
        cuenta = (data.get('cuenta') or data.get('cuentas_balance') or '').strip()
        fecha = (data.get('fecha_renovado') or '').strip()
        def to_float(x):
            try:
                return None if x in (None, '', '-') else float(x)
            except Exception:
                return None
        cupo_asignado = to_float(data.get('cupo_asignado'))
        cupo_ejecutado = to_float(data.get('cupo_ejecutado'))
        garantia = (data.get('garantia') or '').strip() or None
        plazo_meses = data.get('plazo_meses') or data.get('plazo')
        try:
            plazo_meses = int(str(plazo_meses).split()[0]) if plazo_meses not in (None, '') else None
        except Exception:
            plazo_meses = None
        tasa_pct = to_float(data.get('tasa') or data.get('tasa_pct'))

        if not (entidad and fecha and cupo_asignado is not None):
            return Response({'error': 'entidad_financiera, fecha_renovado y cupo_asignado son obligatorios'}, status=400)

        disponible = None
        utilizacion_pct = None
        if cupo_asignado is not None and cupo_ejecutado is not None:
            disponible = cupo_asignado - cupo_ejecutado
            try:
                utilizacion_pct = (cupo_ejecutado / cupo_asignado) * 100 if cupo_asignado else 0
            except Exception:
                utilizacion_pct = None

        with connections['default'].cursor() as c:
            if rec_id:
                c.execute(
                    """
                    UPDATE finanzas.cupos_bancarios
                    SET entidad_financiera=%s, cuentas_balance=%s, fecha_renovado=%s,
                        cupo_asignado=%s, cupo_ejecutado=%s, disponible=COALESCE(%s, cupo_asignado - COALESCE(cupo_ejecutado,0)),
                        garantia=%s, utilizacion_pct=%s, plazo_meses=%s, tasa_pct=%s
                    WHERE id=%s
                    """,
                    [entidad, cuenta or None, fecha, cupo_asignado, cupo_ejecutado, disponible, garantia, utilizacion_pct, plazo_meses, tasa_pct, rec_id]
                )
                return Response({'success': True, 'updated_id': rec_id})
            else:
                c.execute(
                    """
                    INSERT INTO finanzas.cupos_bancarios
                        (entidad_financiera, cuentas_balance, fecha_renovado, cupo_asignado,
                         cupo_ejecutado, disponible, garantia, utilizacion_pct, plazo_meses, tasa_pct, created_at)
                    VALUES (%s,%s,%s,%s,%s,COALESCE(%s, %s - COALESCE(%s,0)),%s,%s,%s,%s, NOW())
                    RETURNING id
                    """,
                    [entidad, cuenta or None, fecha, cupo_asignado, cupo_ejecutado, disponible, cupo_asignado, cupo_ejecutado, garantia, utilizacion_pct, plazo_meses, tasa_pct]
                )
                new_id = c.fetchone()[0]
                return Response({'success': True, 'id': new_id}, status=201)


# ====== Presupuesto ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_presupuesto_spec(request):
    """Esquema de la Tabla 3: Presupuesto."""
    spec = {
        'columns': [
            {'key': 'cuenta', 'label': 'Cuenta (Id_cuenta)', 'type': 'string'},
            {'key': 'nombre_cuenta', 'label': 'Nombre de la Cuenta', 'type': 'string'},
            {'key': 'mes', 'label': 'Mes (1-12)', 'type': 'string'},
            {'key': 'anio', 'label': 'Año (YYYY)', 'type': 'string'},
        ]
    }
    return Response(spec)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def finanzas_presupuesto_list(request):
    """Presupuesto en esquema finanzas (tabla: finanzas.presupuesto).

    GET: lista con join a dim_cuenta para nombre_cuenta. Filtros ?year&month.
    POST: upsert por (cuenta, anio, mes) en finanzas.presupuesto.
    """
    if request.method == 'POST':
        from datetime import datetime as pydt
        payload = request.data
        cuenta = (payload.get('cuenta') or '').strip()
        try:
            anio = int(payload.get('anio'))
            mes = int(payload.get('mes'))
        except Exception:
            return Response({'error': 'anio y mes deben ser enteros'}, status=400)
        try:
            presupuesto = float(payload.get('presupuesto'))
        except Exception:
            return Response({'error': 'presupuesto debe ser numérico'}, status=400)
        if not cuenta:
            return Response({'error': 'cuenta es requerida'}, status=400)

        from django.db import connections
        with connections['default'].cursor() as c:
            # Intentar update primero
            c.execute(
                """
                UPDATE finanzas.presupuesto
                SET presupuesto=%s
                WHERE cuenta=%s AND anio=%s AND mes=%s
                """,
                [presupuesto, cuenta, anio, mes],
            )
            if c.rowcount == 0:
                c.execute(
                    """
                    INSERT INTO finanzas.presupuesto (cuenta, anio, mes, presupuesto, created_at)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [cuenta, anio, mes, presupuesto, pydt.now()],
                )
        return Response({
            'success': True,
            'saved': {
                'cuenta': cuenta,
                'anio': anio,
                'mes': mes,
                'presupuesto': presupuesto,
            }
        }, status=201)

    # GET
    year = request.query_params.get('year')
    month = request.query_params.get('month')
    params = {'year': year, 'month': month}
    where = []
    if year:
        where.append('p.anio = %(year)s::int')
    if month:
        where.append('p.mes = %(month)s::int')
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    sql = f"""
        SELECT p.cuenta::text AS cuenta,
               COALESCE(d.nombre_cuenta, '')::text AS nombre_cuenta,
               p.mes::text AS mes,
               p.anio::text AS anio,
               p.presupuesto::numeric AS presupuesto
        FROM finanzas.presupuesto p
        LEFT JOIN dim_cuenta d ON d.cuenta = p.cuenta
        {where_sql}
        ORDER BY p.anio DESC, p.mes DESC, p.cuenta
    """
    from django.db import connections
    try:
        with connections['default'].cursor() as c:
            c.execute(sql, params)
            cols = [col[0] for col in c.description]
            items = [dict(zip(cols, row)) for row in c.fetchall()]
        return Response({'source': 'finanzas.presupuesto', 'count': len(items), 'items': items, 'filters': {'year': year, 'month': month}})
    except Exception as e:
        return Response({'error': str(e), 'source': 'finanzas.presupuesto'}, status=400)


class PresupuestoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        params = {'year': year, 'month': month}
        where = []
        if year:
            where.append('p.anio = %(year)s::int')
        if month:
            where.append('p.mes = %(month)s::int')
        where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

        sql = f"""
            SELECT p.cuenta::text AS cuenta,
                   COALESCE(d.nombre_cuenta, pc.nombre, '')::text AS nombre_cuenta,
                   p.mes::text AS mes,
                   p.anio::text AS anio,
                   p.presupuesto::numeric AS presupuesto
            FROM finanzas.presupuesto p
            LEFT JOIN dim_cuenta d ON d.cuenta = p.cuenta
            LEFT JOIN plan_cuentas pc ON pc.cuenta = p.cuenta
            {where_sql}
            ORDER BY p.anio DESC, p.mes DESC, p.cuenta
        """
        try:
            with connections['default'].cursor() as c:
                c.execute(sql, params)
                cols = [col[0] for col in c.description]
                items = [dict(zip(cols, row)) for row in c.fetchall()]
            return Response({'source': 'finanzas.presupuesto', 'count': len(items), 'items': items, 'filters': {'year': year, 'month': month}})
        except Exception as e:
            return Response({'error': str(e), 'source': 'finanzas.presupuesto'}, status=400)


class PresupuestoUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Sube un archivo Excel/CSV con columnas: cuenta, anio, mes, presupuesto [, nombre_cuenta]
        y upserta en finanzas.presupuesto.
        """
        f = request.FILES.get('file')
        if not f:
            return Response({'error': "Archivo 'file' requerido"}, status=400)

        filename = getattr(f, 'name', 'upload')
        imported = 0
        errors = []

        import io, csv
        content = f.read()

        def upsert_row(cuenta, anio, mes, presupuesto):
            from datetime import datetime as pydt
            try:
                anio = int(anio)
                mes = int(mes)
                presupuesto = float(str(presupuesto).replace(',', '.'))
            except Exception:
                return False
            with connections['default'].cursor() as c:
                c.execute(
                    """
                    UPDATE finanzas.presupuesto
                    SET presupuesto=%s
                    WHERE cuenta=%s AND anio=%s AND mes=%s
                    """,
                    [presupuesto, str(cuenta), anio, mes],
                )
                if c.rowcount == 0:
                    c.execute(
                        """
                        INSERT INTO finanzas.presupuesto (cuenta, anio, mes, presupuesto, created_at)
                        VALUES (%s,%s,%s,%s,%s)
                        """,
                        [str(cuenta), anio, mes, presupuesto, pydt.now()],
                    )
            return True

        try:
            if filename.lower().endswith('.csv'):
                text = content.decode('utf-8', errors='ignore')
                reader = csv.DictReader(io.StringIO(text))
                for i, row in enumerate(reader, 1):
                    ok = upsert_row(row.get('cuenta'), row.get('anio'), row.get('mes'), row.get('presupuesto'))
                    if ok:
                        imported += 1
                    else:
                        errors.append(f'Fila {i}: datos inválidos')
            else:
                # Intentar XLSX con openpyxl
                try:
                    import openpyxl
                except Exception:
                    return Response({'error': 'Para archivos .xlsx se requiere openpyxl. Sube CSV o instala openpyxl.'}, status=400)
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                # Buscar encabezados
                headers = {}
                for j, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
                    if not cell:
                        continue
                    headers[str(cell).strip().lower()] = j
                required = ['cuenta', 'anio', 'mes', 'presupuesto']
                if not all(k in headers for k in required):
                    return Response({'error': 'Encabezados requeridos: cuenta, anio, mes, presupuesto'}, status=400)
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    values = {k: row[headers[k]] if headers.get(k) is not None else None for k in required}
                    ok = upsert_row(values['cuenta'], values['anio'], values['mes'], values['presupuesto'])
                    if ok: imported += 1
                    else: errors.append(f'Fila {i}: datos inválidos')
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        return Response({'success': True, 'imported': imported, 'errors': errors})

    def post(self, request):
        from datetime import datetime as pydt
        payload = request.data
        cuenta = (payload.get('cuenta') or '').strip()
        try:
            anio = int(payload.get('anio'))
            mes = int(payload.get('mes'))
        except Exception:
            return Response({'error': 'anio y mes deben ser enteros'}, status=400)
        try:
            presupuesto = float(payload.get('presupuesto'))
        except Exception:
            return Response({'error': 'presupuesto debe ser numérico'}, status=400)
        if not cuenta:
            return Response({'error': 'cuenta es requerida'}, status=400)

        with connections['default'].cursor() as c:
            c.execute(
                """
                UPDATE finanzas.presupuesto
                SET presupuesto=%s
                WHERE cuenta=%s AND anio=%s AND mes=%s
                """,
                [presupuesto, cuenta, anio, mes],
            )
            if c.rowcount == 0:
                c.execute(
                    """
                    INSERT INTO finanzas.presupuesto (cuenta, anio, mes, presupuesto, created_at)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [cuenta, anio, mes, presupuesto, pydt.now()],
                )
        return Response({'success': True, 'saved': {'cuenta': cuenta, 'anio': anio, 'mes': mes, 'presupuesto': presupuesto}}, status=201)


# ====== Indicadores Consolidados ======

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finanzas_indicadores_consolidados(request):
    """Indicadores Consolidados desde vistas de BI (prioridad finanzas.vw_indicadores_completos → datamart.vw_indicadores_financieros → finan.vw_oficina_indicadores_mes agregada)."""
    year = int(request.query_params.get('year', datetime.now().year))
    month = int(request.query_params.get('month', datetime.now().month))
    chosen = _choose_view([
        ('finanzas', 'vw_indicadores_completos'),
        ('datamart', 'vw_indicadores_financieros'),
        ('finan', 'vw_oficina_indicadores_mes'),
    ])
    if not chosen:
        return Response({'source': 'none', 'items': []})

    if chosen.endswith('finanzas.vw_indicadores_completos'):
        # Agregamos por nombre_indicador y periodo desde la vista de finanzas
        sql = """
            WITH base AS (
                SELECT nombre_indicador::text,
                       formula::text AS formula_indicador,
                       anio::int, mes::int,
                       AVG(valor_calculado) AS valor
                FROM finanzas.vw_indicadores_completos
                WHERE (anio=%(year)s AND mes=%(month)s)
                   OR (anio=%(year)s-1 AND mes IN (12, %(month)s))
                   OR (anio=%(year)s-2 AND mes=%(month)s)
                   OR (anio=%(year)s-3 AND mes=%(month)s)
                GROUP BY nombre_indicador, formula, anio, mes
            )
            SELECT b0.nombre_indicador,
                   COALESCE(b0.formula_indicador, b1.formula_indicador, b2.formula_indicador, b3.formula_indicador, b4.formula_indicador) AS formula_indicador,
                   b0.valor AS valor_mes_actual,
                   b1.valor AS valor_dic_anterior,
                   b2.valor AS valor_mes_aa,
                   b3.valor AS valor_mes_aa2,
                   b4.valor AS valor_mes_aa3,
                   COALESCE(ii.interpretacion, '')::text AS analisis
            FROM base b0
            LEFT JOIN base b1 ON b1.nombre_indicador=b0.nombre_indicador AND b1.anio=%(year)s-1 AND b1.mes=12
            LEFT JOIN base b2 ON b2.nombre_indicador=b0.nombre_indicador AND b2.anio=%(year)s-1 AND b2.mes=%(month)s
            LEFT JOIN base b3 ON b3.nombre_indicador=b0.nombre_indicador AND b3.anio=%(year)s-2 AND b3.mes=%(month)s
            LEFT JOIN base b4 ON b4.nombre_indicador=b0.nombre_indicador AND b4.anio=%(year)s-3 AND b4.mes=%(month)s
            LEFT JOIN indicadores_interpretacion ii ON ii.indicador=b0.nombre_indicador AND ii.anio=%(year)s AND ii.mes=%(month)s
            WHERE b0.anio=%(year)s AND b0.mes=%(month)s
            ORDER BY b0.nombre_indicador
        """
        params = {'year': year, 'month': month}
    elif chosen.endswith('datamart.vw_indicadores_financieros'):
        sql = """
            SELECT 
                v.nombre_indicador::text,
                v.formula_indicador::text,
                v.valor_indicador_actual::numeric AS valor_mes_actual,
                v.valor_dic_anio_anterior::numeric AS valor_dic_anterior,
                v.valor_mismo_mes_1a::numeric AS valor_mes_aa,
                v.valor_mismo_mes_2a::numeric AS valor_mes_aa2,
                v.valor_mismo_mes_3a::numeric AS valor_mes_aa3,
                COALESCE(ii.interpretacion, '')::text AS analisis
            FROM datamart.vw_indicadores_financieros v
            LEFT JOIN indicadores_interpretacion ii
              ON ii.indicador = v.nombre_indicador AND ii.anio = v.anio AND ii.mes = v.mes
            WHERE v.anio = %(year)s::int AND v.mes = %(month)s::int
            ORDER BY v.nombre_indicador
        """
        params = {'year': year, 'month': month}
    elif chosen.endswith('finan.vw_oficina_indicadores_mes'):
        # Agregar por indicador promedio de valor para el mes, y construir históricos relativos
        sql = """
            WITH base AS (
                SELECT indicador::text AS nombre_indicador,
                       %(year)s::int AS anio, %(month)s::int AS mes,
                       AVG(CASE WHEN anio=%(year)s AND mes=%(month)s THEN valor END) AS valor_mes_actual,
                       AVG(CASE WHEN anio=%(year)s-1 AND mes=12 THEN valor END) AS valor_dic_anterior,
                       AVG(CASE WHEN anio=%(year)s-1 AND mes=%(month)s THEN valor END) AS valor_mes_aa,
                       AVG(CASE WHEN anio=%(year)s-2 AND mes=%(month)s THEN valor END) AS valor_mes_aa2,
                       AVG(CASE WHEN anio=%(year)s-3 AND mes=%(month)s THEN valor END) AS valor_mes_aa3
                FROM finan.vw_oficina_indicadores_mes
                WHERE (anio=%(year)s AND mes=%(month)s)
                   OR (anio=%(year)s-1 AND mes IN (12, %(month)s))
                   OR (anio=%(year)s-2 AND mes=%(month)s)
                   OR (anio=%(year)s-3 AND mes=%(month)s)
                GROUP BY indicador
            )
            SELECT b.nombre_indicador,
                   f.formula_indicador::text,
                   b.valor_mes_actual::numeric,
                   b.valor_dic_anterior::numeric,
                   b.valor_mes_aa::numeric,
                   b.valor_mes_aa2::numeric,
                   b.valor_mes_aa3::numeric,
                   COALESCE(ii.interpretacion, '')::text AS analisis
            FROM base b
            LEFT JOIN indicadores_financieros f ON f.nombre_indicador = b.nombre_indicador
            LEFT JOIN indicadores_interpretacion ii ON ii.indicador = b.nombre_indicador AND ii.anio=%(year)s AND ii.mes=%(month)s
            ORDER BY b.nombre_indicador
        """
        params = {'year': year, 'month': month}
    else:
        # Caso genérico para finanzas.vw_indicadores_completos con nombres esperados
        sql = """
            SELECT nombre_indicador::text,
                   formula_indicador::text,
                   valor_mes_actual::numeric,
                   valor_dic_anterior::numeric,
                   valor_mes_aa::numeric,
                   valor_mes_aa2::numeric,
                   valor_mes_aa3::numeric,
                   COALESCE(analisis, '')::text AS analisis
            FROM finanzas.vw_indicadores_completos
            WHERE anio=%(year)s::int AND mes=%(month)s::int
            ORDER BY nombre_indicador
        """
        params = {'year': year, 'month': month}

    with connections['default'].cursor() as c:
        c.execute(sql, params)
        cols = [col[0] for col in c.description]
        items = [dict(zip(cols, row)) for row in c.fetchall()]
    return Response({'year': year, 'month': month, 'items': items, 'source': chosen})


class IndicadoresAnalisisView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Guarda/actualiza el análisis para un indicador en (anio, mes)."""
        data = request.data
        indicador = data.get('indicador') or data.get('nombre_indicador')
        try:
            anio = int(data.get('anio'))
            mes = int(data.get('mes'))
        except Exception:
            return Response({'error': 'anio y mes deben ser enteros'}, status=400)
        analisis = (data.get('analisis') or '').strip()
        if not indicador:
            return Response({'error': 'indicador es requerido'}, status=400)
        with connections['default'].cursor() as c:
            c.execute(
                """
                UPDATE indicadores_interpretacion
                SET interpretacion=%s, updated_at=NOW()
                WHERE indicador=%s AND anio=%s AND mes=%s
                """,
                [analisis, indicador, anio, mes],
            )
            if c.rowcount == 0:
                c.execute(
                    """
                    INSERT INTO indicadores_interpretacion (indicador, anio, mes, interpretacion, updated_at)
                    VALUES (%s,%s,%s,%s,NOW())
                    """,
                    [indicador, anio, mes, analisis],
                )
        return Response({'success': True})
